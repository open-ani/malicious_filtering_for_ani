import json

import pandas as pd
import numpy as np
import torch
from transformers import DistilBertTokenizer, DistilBertForSequenceClassification, Trainer, TrainingArguments
from datasets import Dataset
from sklearn.metrics import accuracy_score, precision_recall_fscore_support
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import psutil


def load_test_data(data_path, tokenizer):
    df = pd.read_csv(data_path, sep=',', header=0)
    df = df[['label', 'text']]
    data_labels = df['label'].tolist()
    data_text = df['text'].tolist()

    # Tokenize the datasets
    test_encodings = tokenizer(data_text, truncation=True, padding=True, max_length=512)

    test_dataset = Dataset.from_dict({
        'input_ids': test_encodings['input_ids'],
        'attention_mask': test_encodings['attention_mask'],
        'labels': data_labels
    })

    return test_dataset, df


def compute_metrics(pred):
    labels = pred.label_ids
    preds = pred.predictions.argmax(-1)
    accuracy = accuracy_score(labels, preds)
    precision, recall, f1, _ = precision_recall_fscore_support(labels, preds, average='binary')
    return {
        'accuracy': accuracy,
        'precision': precision,
        'recall': recall,
        'f1': f1,
    }


def evaluate_model_and_run_example():
    # Load the fine-tuned model and tokenizer
    save_directory = './model_small'
    tokenizer = DistilBertTokenizer.from_pretrained(save_directory)
    model = DistilBertForSequenceClassification.from_pretrained(save_directory)

    # Load test data
    test_data_path = '../../data/toxic_comment_data/toxic_comment_test.csv'
    test_dataset, df = load_test_data(test_data_path, tokenizer)

    # Create Trainer instance for prediction
    training_args = TrainingArguments(
        output_dir='./results_small',  # output directory
        per_device_eval_batch_size=16,  # batch size for evaluation
    )

    trainer = Trainer(
        model=model,
        args=training_args,
    )

    # Track memory usage before prediction
    process = psutil.Process()
    memory_before = process.memory_info().rss / 1024 ** 2  # in MB
    print(f"Memory usage before prediction: {memory_before:.2f} MB")

    # Predict on test data
    predictions = trainer.predict(test_dataset)

    # Track memory usage after prediction
    memory_after = process.memory_info().rss / 1024 ** 2  # in MB
    print(f"Memory usage after prediction: {memory_after:.2f} MB")

    logits = predictions.predictions

    # Get the class with the highest predicted probability
    preds = logits.argmax(-1)

    # Calculate the softmax to get probabilities
    probabilities = np.exp(logits) / np.exp(logits).sum(-1, keepdims=True)

    # Get the confidence level for each prediction
    confidences = probabilities.max(-1)

    df['confidence'] = confidences

    # Add predictions to the DataFrame
    df['predictions'] = preds

    # Highlight incorrect predictions
    df['correct'] = df['label'] == df['predictions']

    # calculate sum of incorrect confidence

    incorrect_confidence = df[df['correct'] == False]['confidence'].sum()

    print(f"Sum of incorrect confidence: {incorrect_confidence}")

    # Save to Excel
    save_to_excel(df, 'test_results_2.xlsx')

    # Calculate evaluation metrics
    labels = df['label'].tolist()
    accuracy = accuracy_score(labels, preds)
    precision, recall, f1, _ = precision_recall_fscore_support(labels, preds, average='binary')

    print(f"Accuracy: {accuracy}")
    print(f"Precision: {precision}")
    print(f"Recall: {recall}")
    print(f"F1 Score: {f1}")


def save_to_excel(df, file_name):
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Results"

    # Create header
    headers = list(df.columns)
    ws.append(headers)

    # Define fill for incorrect predictions
    fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

    # Add data
    for index, row in df.iterrows():
        ws.append(list(row))
        if not row['correct']:
            for cell in ws[index + 2]:  # +2 because Excel is 1-indexed and we have a header row
                cell.fill = fill

    wb.save(file_name)


if __name__ == '__main__':
    evaluate_model_and_run_example()
